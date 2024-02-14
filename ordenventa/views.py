from rest_framework import generics
from .models import OrdenVenta, ItemOrdenVenta
from .serializers import OrdenVentaSerializer, ItemOrdenVentaSerializer
from django.shortcuts import redirect, render, get_object_or_404
from django.views import View
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect, JsonResponse
from .forms import ItemOrdenVentaForm, OrdenVentaForm
from openpyxl import load_workbook
import openpyxl
import csv


class OrdenVentaCRUDView(View):
    template_name = "ordenventa_crud.html"

    def get(self, request, *args, **kwargs):
        # Obtén la acción de la URL (edit o delete)
        action = request.GET.get("action", None)

        # Lógica para editar
        if action == "edit":
            orden_venta_id = request.GET.get("edit", None)
            if orden_venta_id:
                orden_venta = get_object_or_404(OrdenVenta, pk=orden_venta_id)
                form = OrdenVentaForm(instance=orden_venta)
            else:
                form = OrdenVentaForm()

        # Lógica para eliminar
        elif action == "delete":
            orden_venta_id = request.GET.get("delete", None)
            if orden_venta_id:
                orden_venta = get_object_or_404(OrdenVenta, pk=orden_venta_id)
                orden_venta.delete()
                return HttpResponseRedirect(reverse_lazy("ordenventa-crud"))

        # Lógica por defecto para mostrar la lista y el formulario de creación
        else:
            form = OrdenVentaForm()

        ordenes_venta = OrdenVenta.objects.all()
        return render(
            request, self.template_name, {"ordenes_venta": ordenes_venta, "form": form}
        )

    def post(self, request, *args, **kwargs):
        form = OrdenVentaForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse_lazy("ordenventa-crud"))
        else:
            ordenes_venta = OrdenVenta.objects.all()
            return render(
                request,
                self.template_name,
                {"ordenes_venta": ordenes_venta, "form": form},
            )


def agregar_item_orden_venta(request, ordenventa_id):
    # Obtén la instancia de OrdenVenta
    orden_venta = OrdenVenta.objects.get(pk=ordenventa_id)

    if request.method == "POST":
        form = ItemOrdenVentaForm(request.POST)
        if form.is_valid():
            form.save(commit=False)

            # Procesar archivo Excel
            if "archivo_excel" in request.FILES:
                archivo_excel = request.FILES["archivo_excel"]
                procesar_datos_desde_excel(archivo_excel, orden_venta)

            # Continuar con el procesamiento del formulario y la redirección
            form.save()

            # Puedes redirigir a otra página o hacer lo que necesites después de guardar
            return redirect("nombre_de_la_vista")
    else:
        form = ItemOrdenVentaForm()

    return render(
        request, "tu_template.html", {"form": form, "orden_venta": orden_venta}
    )


def procesar_datos_desde_excel(archivo_excel, orden_venta):
    # Abre el archivo Excel
    wb = openpyxl.load_workbook(archivo_excel)
    sheet = wb.active

    # Itera sobre las filas del archivo Excel (empezando desde la segunda fila, suponiendo que la primera fila es encabezado)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Crea una instancia de ItemOrdenVenta con los datos de la fila
        item_orden_venta = ItemOrdenVenta(
            ordenventa=orden_venta,
            nro_articulo=row[0],
            cantidad=row[1],
            precio_bruto=row[2],
            total_bruto=row[3],
        )

        # Guarda el nuevo item en la base de datos
        item_orden_venta.save()


# Vistas para OrdenVenta
class ListaOrdenesVenta(generics.ListCreateAPIView):
    queryset = OrdenVenta.objects.all()
    serializer_class = OrdenVentaSerializer


class DetalleOrdenVenta(generics.RetrieveUpdateDestroyAPIView):
    queryset = OrdenVenta.objects.all()
    serializer_class = OrdenVentaSerializer


# Vistas para ItemOrdenVenta
class ListaItemsOrdenVenta(generics.ListCreateAPIView):
    queryset = ItemOrdenVenta.objects.all()
    serializer_class = ItemOrdenVentaSerializer


class DetalleItemOrdenVenta(generics.RetrieveUpdateDestroyAPIView):
    queryset = ItemOrdenVenta.objects.all()
    serializer_class = ItemOrdenVentaSerializer
